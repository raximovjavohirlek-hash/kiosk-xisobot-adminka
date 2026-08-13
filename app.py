import os
import sys
import json
import time
import webbrowser
import threading
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = 'kiosk-xisobot-secret-key-2026'
app.config['UPLOAD_FOLDER'] = os.path.dirname(os.path.abspath(__file__))
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB limit
app.config['ADMIN_PASSWORD'] = os.environ.get('ADMIN_PASSWORD', 'admin')

MAPPINGS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'kiosk_mappings.json')
UPLOAD_LOGS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'kiosk_upload_logs.json')

DEFAULT_EMAIL_MAP = {
    "toshkent.shimoliykiosk@railway.uz": {"station": "Тошкент Марказий", "col_soni": 30, "col_summa": 31},
    "kiosk@axonlogic.uz": {"station": "Тошкент Жанубий", "col_soni": 10, "col_summa": 11},
    "samarqandkiosk@railway.uz": {"station": "Самарқанд", "col_soni": 26, "col_summa": 27},
    "urganchkiosk@railway.uz": {"station": "Урганч", "col_soni": 32, "col_summa": 33},
    "khivakiosk@railway.uz": {"station": "Хива", "col_soni": 8, "col_summa": 9},
    "navoiykiosk@railway.uz": {"station": "Навои", "col_soni": 14, "col_summa": 15},
    "buxorokiosk@railway.uz": {"station": "Бухоро", "col_soni": 6, "col_summa": 7},
    "qongirotkiosk@railway.uz": {"station": "Қўнғирод", "col_soni": 22, "col_summa": 23},
    "nukuskiosk@railway.uz": {"station": "Нукус", "col_soni": 18, "col_summa": 19},
    "andijonkiosk@railway.uz": {"station": "Андижон", "col_soni": 4, "col_summa": 5},
    "qoqonkiosk@railway.uz": {"station": "Қўқон", "col_soni": 24, "col_summa": 25},
    "margilonkiosk@railway.uz": {"station": "Марғилон", "col_soni": 12, "col_summa": 13},
    "namangankiosk@railway.uz": {"station": "Наманган", "col_soni": 16, "col_summa": 17},
    "termizkiosk@railway.uz": {"station": "Термиз", "col_soni": 28, "col_summa": 29},
    "qarshikiosk@railway.uz": {"station": "Қарши", "col_soni": 20, "col_summa": 21}
}

ONLINE_PAYMENTS = ['HamkorbankHold', 'HamkorbankWebView', 'Payme', 'StripeIntegration', 'OctoBankFC']

def load_mappings():
    if os.path.exists(MAPPINGS_FILE):
        try:
            with open(MAPPINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return DEFAULT_EMAIL_MAP

def save_mappings(mappings):
    with open(MAPPINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)

def load_upload_logs():
    if os.path.exists(UPLOAD_LOGS_FILE):
        try:
            with open(UPLOAD_LOGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def add_upload_log(filename, rows_count, status="Muvaffaqiyatli"):
    logs = load_upload_logs()
    logs.insert(0, {
        "id": len(logs) + 1,
        "filename": filename,
        "rows": rows_count,
        "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        "status": status
    })
    logs = logs[:50]
    with open(UPLOAD_LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

MONTH_NAMES_UZ = {
    1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
    5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
    9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
}

def parse_report_file(fpath, fname):
    email_map = load_mappings()
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        if 'Худудлар' not in wb.sheetnames:
            return None, None
        ws_h = wb['Худудлар']
        ws_j = wb['Жами'] if 'Жами' in wb.sheetnames else None

        fn_lower = fname.lower()
        if 'январ' in fn_lower: ym_code = '2026-01'
        elif 'феврал' in fn_lower: ym_code = '2026-02'
        elif 'март' in fn_lower: ym_code = '2026-03'
        elif 'апрел' in fn_lower: ym_code = '2026-04'
        elif 'маи' in fn_lower or 'май' in fn_lower: ym_code = '2026-05'
        elif 'июн' in fn_lower: ym_code = '2026-06'
        elif 'июл' in fn_lower: ym_code = '2026-07'
        elif 'август' in fn_lower: ym_code = '2026-08'
        else:
            return None, None

        station_totals = {email: {'tickets': 0, 'summa': 0, 'daily': []} for email in email_map}
        daily_trend = []

        for r in range(4, 35):
            row_date = ws_h.cell(r, 1).value
            if not row_date:
                break
            if isinstance(row_date, str):
                s_val = row_date.strip().lower()
                if 'жам' in s_val or 'jam' in s_val or 'total' in s_val:
                    break
            d_str = row_date.strftime('%d.%m.%Y') if hasattr(row_date, 'strftime') else str(row_date)

            day_tickets = 0
            day_summa = 0
            for email, meta in email_map.items():
                c_soni = meta['col_soni']
                c_summa = meta['col_summa']
                v_soni = ws_h.cell(r, c_soni).value or 0
                v_summa = ws_h.cell(r, c_summa).value or 0
                try: v_soni = int(v_soni); v_summa = int(v_summa)
                except Exception: v_soni, v_summa = 0, 0

                station_totals[email]['tickets'] += v_soni
                station_totals[email]['summa'] += v_summa
                station_totals[email]['daily'].append({
                    'date': d_str,
                    'tickets': v_soni,
                    'summa': v_summa
                })
                day_tickets += v_soni
                day_summa += v_summa

            raw_on_t = ws_j.cell(r, 4).value if ws_j else 0
            raw_on_s = ws_j.cell(r, 5).value if ws_j else 0
            raw_term_t = ws_j.cell(r, 6).value if ws_j else 0
            raw_term_s = ws_j.cell(r, 7).value if ws_j else 0

            term_t = int(raw_term_t) if isinstance(raw_term_t, (int, float)) else 0
            term_s = int(raw_term_s) if isinstance(raw_term_s, (int, float)) else 0
            on_t = int(raw_on_t) if isinstance(raw_on_t, (int, float)) else max(0, day_tickets - term_t)
            on_s = int(raw_on_s) if isinstance(raw_on_s, (int, float)) else max(0, day_summa - term_s)

            daily_trend.append({
                'date': d_str, 'tickets': day_tickets, 'summa': day_summa,
                'online_tickets': on_t, 'online_summa': on_s,
                'terminal_tickets': term_t, 'terminal_summa': term_s
            })

        station_sums = []
        m_total_tickets = sum(station_totals[email]['tickets'] for email in email_map)
        m_total_summa = sum(station_totals[email]['summa'] for email in email_map)

        for email, meta in email_map.items():
            st_name = meta['station']
            s_val = station_totals[email]['summa']
            t_val = station_totals[email]['tickets']
            sh_pct = round((s_val / m_total_summa) * 100, 1) if m_total_summa > 0 else 0
            station_sums.append({
                'stansiya': st_name, 'email': email,
                'soni_val': t_val,
                'summa_val': s_val,
                'share_percent': sh_pct,
                'daily_breakdown': station_totals[email]['daily']
            })

        by_summa = sorted(station_sums, key=lambda x: x['summa_val'], reverse=True)

        return ym_code, {
            'total_tickets': m_total_tickets, 'total_summa': m_total_summa,
            'stations': by_summa, 'daily_trend': daily_trend
        }
    except Exception as ex:
        print("parse_report_file error:", fname, ex)
        return None, None

def get_all_official_monthly_reports():
    reports = {}
    search_paths = [
        app.config['UPLOAD_FOLDER'],
        os.path.join(app.config['UPLOAD_FOLDER'], 'excellar')
    ]
    for sp in search_paths:
        if os.path.exists(sp):
            for fn in sorted(os.listdir(sp)):
                if fn.endswith('.xlsx') and not fn.startswith('~$'):
                    fp = os.path.join(sp, fn)
                    ym, stats = parse_report_file(fp, fn)
                    if ym and stats and ym not in reports:
                        reports[ym] = stats
    return reports

def process_excel(data_path, report_path):
    email_map = load_mappings()
    
    # Automatically update official report workbook if raw data.xlsx exists
    if data_path and os.path.exists(data_path):
        try:
            from update_august_report import update_kiosk_report
            update_kiosk_report(data_path, report_path)
        except Exception as ex:
            print("auto update_kiosk_report warning:", ex)
    
    # 1. Load official monthly excel reports as primary source of truth
    monthly_data = get_all_official_monthly_reports()

    # 2. If raw data.xlsx exists and has months not in official reports, add them
    if data_path and os.path.exists(data_path):
        try:
            df = pd.read_excel(data_path)
            if 'Дата создания' in df.columns:
                df['Дата создания_dt'] = pd.to_datetime(df['Дата создания'], errors='coerce')
                df['Date'] = df['Дата создания_dt'].dt.date
                df['YearMonth'] = df['Дата создания_dt'].dt.strftime('%Y-%m')
                
                kiosk_df = df[df['Пользователь'].isin(email_map.keys())].copy()
                kiosk_df['PaymentType'] = kiosk_df['Способ оплаты'].apply(
                    lambda x: 'Online' if x in ONLINE_PAYMENTS else 'Terminal'
                )
                
                unique_periods = sorted([p for p in kiosk_df['YearMonth'].dropna().unique()], reverse=True)
                for ym in unique_periods:
                    if ym not in monthly_data: # Only add if not present in official monthly reports!
                        m_kiosk_df = kiosk_df[kiosk_df['YearMonth'] == ym]
                        
                        m_grouped_station = m_kiosk_df.groupby(['Date', 'Пользователь']).agg(
                            tickets=('Количество билетов', 'sum'),
                            summa=('Общая стоимость', 'sum')
                        ).reset_index()
                        
                        m_grouped_payment = m_kiosk_df.groupby(['Date', 'PaymentType']).agg(
                            tickets=('Количество билетов', 'sum'),
                            summa=('Общая стоимость', 'sum')
                        ).reset_index()
                        
                        m_station_sums = []
                        for email, meta in email_map.items():
                            st_name = meta['station']
                            m_match = m_grouped_station[m_grouped_station['Пользователь'] == email]
                            soni_val = int(m_match['tickets'].sum()) if not m_match.empty else 0
                            summa_val = int(m_match['summa'].sum()) if not m_match.empty else 0
                            m_station_sums.append({
                                'stansiya': st_name,
                                'email': email,
                                'soni_val': soni_val,
                                'summa_val': summa_val
                            })
                            
                        m_by_summa = sorted(m_station_sums, key=lambda x: x['summa_val'], reverse=True)
                        m_total_tickets = sum(s['soni_val'] for s in m_station_sums)
                        m_total_summa = sum(s['summa_val'] for s in m_station_sums)
                        
                        m_daily_trend = []
                        all_dates_in_m = sorted(m_kiosk_df['Date'].dropna().unique())
                        for d in all_dates_in_m:
                            d_str = d.strftime('%d.%m.%Y')
                            d_st_df = m_grouped_station[m_grouped_station['Date'] == d]
                            d_tickets = int(d_st_df['tickets'].sum())
                            d_summa = int(d_st_df['summa'].sum())
                            
                            d_pay_df = m_grouped_payment[m_grouped_payment['Date'] == d]
                            on_row = d_pay_df[d_pay_df['PaymentType'] == 'Online']
                            term_row = d_pay_df[d_pay_df['PaymentType'] == 'Terminal']
                            
                            on_t = int(on_row['tickets'].values[0]) if not on_row.empty else 0
                            on_s = int(on_row['summa'].values[0]) if not on_row.empty else 0
                            term_t = int(term_row['tickets'].values[0]) if not term_row.empty else 0
                            term_s = int(term_row['summa'].values[0]) if not term_row.empty else 0
                            
                            m_daily_trend.append({
                                'date': d_str,
                                'tickets': d_tickets,
                                'summa': d_summa,
                                'online_tickets': on_t,
                                'online_summa': on_s,
                                'terminal_tickets': term_t,
                                'terminal_summa': term_s
                            })
                            
                        monthly_data[ym] = {
                            'total_tickets': m_total_tickets,
                            'total_summa': m_total_summa,
                            'stations': m_by_summa,
                            'daily_trend': m_daily_trend
                        }
        except Exception as ex:
            print("data_path parse warning:", ex)

    # Build available_months sorted descending
    all_ym_codes = sorted(list(monthly_data.keys()), reverse=True)
    available_months = []
    for ym in all_ym_codes:
        try:
            y, m = ym.split('-')
            m_name = MONTH_NAMES_UZ.get(int(m), ym)
            available_months.append({
                'code': ym,
                'name': f"{m_name} {y}"
            })
        except Exception:
            available_months.append({'code': ym, 'name': ym})

    # All-time / overall totals across all months in monthly_data
    overall_station_totals = {email: {'tickets': 0, 'summa': 0} for email in email_map}
    overall_daily_trend = []
    
    for ym, m_info in monthly_data.items():
        for st in (m_info.get('stations') or []):
            em = st.get('email')
            if em in overall_station_totals:
                overall_station_totals[em]['tickets'] += st.get('soni_val', 0)
                overall_station_totals[em]['summa'] += st.get('summa_val', 0)
        overall_daily_trend.extend(m_info.get('daily_trend') or [])

    all_station_sums = []
    for email, meta in email_map.items():
        st_name = meta['station']
        all_station_sums.append({
            'stansiya': st_name,
            'email': email,
            'soni_val': overall_station_totals[email]['tickets'],
            'summa_val': overall_station_totals[email]['summa']
        })

    overall_by_summa = sorted(all_station_sums, key=lambda x: x['summa_val'], reverse=True)
    overall_total_tickets = sum(s['soni_val'] for s in all_station_sums)
    overall_total_summa = sum(s['summa_val'] for s in all_station_sums)

    overall_data_map = {
        'total_tickets': overall_total_tickets,
        'total_summa': overall_total_summa,
        'stations': overall_by_summa,
        'daily_trend': overall_daily_trend
    }

    return enrich_stats_with_executive_metrics(monthly_data, overall_data_map, available_months)

def enrich_stats_with_executive_metrics(monthly_data_map, overall_data_map, available_months):
    def format_stations(station_list, total_summa):
        enriched = []
        for st in station_list:
            soni = st.get('soni_val', 0)
            summa = st.get('summa_val', 0)
            avg_price = round(summa / soni) if soni > 0 else 0
            share_pct = round((summa / total_summa * 100), 1) if total_summa > 0 else 0.0
            
            st_copy = dict(st)
            st_copy.update({
                'avg_price': avg_price,
                'share_percent': share_pct
            })
            enriched.append(st_copy)
        return enriched

    for ym, m_info in monthly_data_map.items():
        t_sum = m_info.get('total_summa', 0)
        m_info['stations'] = format_stations(m_info.get('stations', []), t_sum)

    tot_sum = overall_data_map.get('total_summa', 0)
    overall_data_map['stations'] = format_stations(overall_data_map.get('stations', []), tot_sum)

    latest_ym = available_months[0]['code'] if available_months else None
    default_stats = monthly_data_map.get(latest_ym) if latest_ym else overall_data_map

    # Real Executive Summary Calculations
    total_summa = default_stats.get('total_summa', 0)
    total_tickets = default_stats.get('total_tickets', 0)
    overall_avg_price = round(total_summa / total_tickets) if total_tickets > 0 else 0
    
    daily_trend = default_stats.get('daily_trend', [])
    trend_len = len(daily_trend)
    daily_avg_summa = round(total_summa / trend_len) if trend_len > 0 else 0
    daily_avg_tickets = round(total_tickets / trend_len) if trend_len > 0 else 0

    # Peak Revenue Date
    peak_day = max(daily_trend, key=lambda x: x.get('summa', 0)) if daily_trend else {'date': '-', 'summa': 0, 'tickets': 0}
    
    stations_list = default_stats.get('stations', [])
    top_station = stations_list[0] if stations_list else {'stansiya': "Noma'lum", 'summa_val': 0, 'soni_val': 0, 'share_percent': 0}
    second_station = stations_list[1] if len(stations_list) > 1 else {'stansiya': "-", 'summa_val': 0, 'share_percent': 0}

    # Online vs Terminal ratio
    online_tickets = sum(d.get('online_tickets', 0) for d in daily_trend)
    terminal_tickets = sum(d.get('terminal_tickets', 0) for d in daily_trend)
    total_pay_tickets = online_tickets + terminal_tickets or 1
    online_pct = round((online_tickets / total_pay_tickets) * 100, 1)

    executive_summary = {
        'net_revenue': total_summa,
        'total_tickets': total_tickets,
        'overall_avg_price': overall_avg_price,
        'daily_avg_revenue': daily_avg_summa,
        'daily_avg_tickets': daily_avg_tickets,
        'peak_date': peak_day.get('date', '-'),
        'peak_day_revenue': peak_day.get('summa', 0),
        'top_station': top_station.get('stansiya'),
        'top_station_summa': top_station.get('summa_val', 0),
        'top_station_share': top_station.get('share_percent', 0),
        'second_station': second_station.get('stansiya'),
        'second_station_summa': second_station.get('summa_val', 0),
        'online_percent': online_pct,
        'terminal_percent': round(100 - online_pct, 1) if online_pct else 0.0,
        'ai_recommendation': f"Hurmatli Rahbariyat, ushbu davrda kiosklar bo'yicha jami {total_summa:,} so'm tushum hamda {total_tickets:,} ta chipta sotildi. "
                            f"Bitta chiptaning o'rtacha narxi {overall_avg_price:,} so'mni va kunlik o'rtacha tushum {daily_avg_summa:,} so'mni tashkil etdi. "
                            f"Eng savdoli kassa {top_station.get('stansiya')} bo'lib, uning umumiy tushumdagi ulushi {top_station.get('share_percent')}% ni tashkil qiladi. "
                            f"Eng yuqori savdo ko'rsatkichi {peak_day.get('date')} sanasida ({peak_day.get('summa', 0):,} so'm) qayd etilgan."
    }

    return {
        'total_tickets': default_stats['total_tickets'],
        'total_summa': default_stats['total_summa'],
        'stations': default_stats['stations'],
        'daily_trend': default_stats['daily_trend'],
        'available_months': available_months,
        'monthly_data': monthly_data_map,
        'overall_data': overall_data_map,
        'director_summary': executive_summary,
        'last_updated': datetime.now().strftime('%d.%m.%Y %H:%M')
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/stats', methods=['GET'])
def get_stats():
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
    try:
        stats = process_excel(data_path, report_path)
        all_reports = get_all_official_monthly_reports()
        return jsonify({'success': True, 'stats': stats, 'monthly_reports': all_reports})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Fayl tanlanmagan'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Fayl tanlanmagan'}), 400

    filename = secure_filename(file.filename)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        file.save(save_path)
        # If it's a monthly report file, also copy to excellar/
        if 'кисока' in filename.lower() or 'киоска' in filename.lower():
            ex_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'excellar')
            os.makedirs(ex_dir, exist_ok=True)
            import shutil
            shutil.copy(save_path, os.path.join(ex_dir, filename))

        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
        data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        stats = process_excel(data_path, report_path)
        
        add_upload_log(file.filename, 1, "Muvaffaqiyatli")
        
        return jsonify({
            'success': True, 
            'message': "Fayl muvaffaqiyatli yuklandi!", 
            'stats': stats
        })
    except Exception as e:
        add_upload_log(file.filename, 0, f"Xatolik: {str(e)}")
        return jsonify({'success': False, 'error': f"Xatolik yuz berdi: {str(e)}"}), 500

@app.route('/api/download', methods=['GET'])
def download():
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    if os.path.exists(report_path):
        return send_file(report_path, as_attachment=True, download_name='Август_киоска_hisobot.xlsx')
    return jsonify({'error': 'Fayl topilmadi'}), 404

@app.route('/api/export-station-excel/<path:station_name>', methods=['GET'])
def export_station_excel(station_name):
    try:
        requested_month = request.args.get('month')
        all_reports = get_all_official_monthly_reports()
        
        if requested_month and requested_month in all_reports:
            stats = all_reports[requested_month]
            m_code_str = requested_month
        else:
            report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
            data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
            stats = process_excel(data_path, report_path)
            m_code_str = '2026-08'
        
        stations = stats.get('stations', [])
        station_data = None
        rank = 0
        for idx, s in enumerate(stations):
            if s.get('stansiya') == station_name:
                station_data = s
                rank = idx + 1
                break
        
        if not station_data:
            return jsonify({'error': 'Stansiya topilmadi'}), 404

        # Create openpyxl workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Hisobot - {station_name[:20]}"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name='Calibri', size=15, bold=True, color='1F2937')
        subtitle_font = Font(name='Calibri', size=11, italic=True, color='4B5563')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Deep Blue
        
        kpi_title_font = Font(name='Calibri', size=10, bold=True, color='4B5563')
        kpi_val_font = Font(name='Calibri', size=13, bold=True, color='1E3A8A')
        kpi_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        
        total_font = Font(name='Calibri', size=11, bold=True, color='065F46')
        total_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid') # Emerald
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Title Block
        ws.merge_cells('A1:E1')
        ws['A1'] = f"O'ZBEKISTON TEMIR YO'LLARI — KIOSK HISOBOTI"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_left

        ws.merge_cells('A2:E2')
        ws['A2'] = f"Kassa / Stansiya: {station_data['stansiya']} ({rank}-O'rin) | Shakllangan vaqt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = align_left

        # KPI Cards Block (Rows 4-5)
        ws['A4'] = "Jami Tushum Summasi"
        ws['A5'] = f"{station_data['summa_val']:,} so'm"
        
        ws['B4'] = "Sotilgan Chiptalar"
        ws['B5'] = f"{station_data['soni_val']:,} ta"
        
        avg_p = round(station_data['summa_val'] / station_data['soni_val']) if station_data['soni_val'] > 0 else 0
        ws['C4'] = "O'rtacha Chek"
        ws['C5'] = f"{avg_p:,} so'm"
        
        ws['D4'] = "Tushum Ulushi"
        ws['D5'] = f"{station_data.get('share_percent', 0)}%"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = kpi_title_font
            ws[f'{col}4'].fill = kpi_fill
            ws[f'{col}4'].alignment = align_center
            ws[f'{col}5'].font = kpi_val_font
            ws[f'{col}5'].fill = kpi_fill
            ws[f'{col}5'].alignment = align_center

        # Table Headers (Row 7)
        headers = ["№", "Sana", "Sotilgan Chiptalar (ta)", "Kunlik Tushum (so'm)", "O'rtacha Chek (so'm)"]
        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # Table Data (Row 8+)
        daily_list = station_data.get('daily_breakdown', [])
        row_idx = 8
        total_tickets = 0
        total_sum = 0

        for i, day in enumerate(daily_list, 1):
            t_val = day.get('tickets', 0)
            s_val = day.get('summa', 0)
            day_avg = round(s_val / t_val) if t_val > 0 else 0
            
            total_tickets += t_val
            total_sum += s_val

            ws.cell(row=row_idx, column=1, value=i).alignment = align_center
            ws.cell(row=row_idx, column=2, value=day.get('date', '')).alignment = align_center
            
            c3 = ws.cell(row=row_idx, column=3, value=t_val)
            c3.alignment = align_right
            c3.number_format = '#,##0'
            
            c4 = ws.cell(row=row_idx, column=4, value=s_val)
            c4.alignment = align_right
            c4.number_format = '#,##0" so\'m"'
            
            c5 = ws.cell(row=row_idx, column=5, value=day_avg)
            c5.alignment = align_right
            c5.number_format = '#,##0" so\'m"'

            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).border = thin_border

            row_idx += 1

        # Total Row
        grand_avg = round(total_sum / total_tickets) if total_tickets > 0 else 0
        ws.cell(row=row_idx, column=1, value="—").alignment = align_center
        
        c2 = ws.cell(row=row_idx, column=2, value="JAMI (YAKUNIY)")
        c2.font = total_font
        c2.alignment = align_left
        
        c3 = ws.cell(row=row_idx, column=3, value=total_tickets)
        c3.font = total_font
        c3.alignment = align_right
        c3.number_format = '#,##0'

        c4 = ws.cell(row=row_idx, column=4, value=total_sum)
        c4.font = total_font
        c4.alignment = align_right
        c4.number_format = '#,##0" so\'m"'

        c5 = ws.cell(row=row_idx, column=5, value=grand_avg)
        c5.font = total_font
        c5.alignment = align_right
        c5.number_format = '#,##0" so\'m"'

        for c in range(1, 6):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = total_fill
            cell.border = thin_border

        # Adjust Column Widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_st_name = "".join(c for c in station_name if c.isalnum() or c in (' ', '_', '-')).strip()
        download_name = f"Kiosk_Hisobot_{safe_st_name}_Avgust_2026.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print("export_station_excel error:", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/mappings', methods=['GET', 'POST'])
def handle_mappings():
    if request.method == 'POST':
        new_map = request.json
        save_mappings(new_map)
        return jsonify({'success': True, 'message': 'Pochta biriktirmalari saqlandi!'})
    return jsonify({'success': True, 'mappings': load_mappings()})

@app.route('/api/upload-logs', methods=['GET'])
def get_upload_logs():
    return jsonify({'success': True, 'logs': load_upload_logs()})

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json or {}
    password = data.get('password', '')
    if password == app.config['ADMIN_PASSWORD']:
        return jsonify({'success': True, 'message': 'Admin rejimiga kirdingiz!'})
    return jsonify({'success': False, 'error': "Parol noto'g'ri!"}), 401

@app.route('/api/director-summary', methods=['GET'])
def get_director_summary():
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
    try:
        stats = process_excel(data_path, report_path)
        return jsonify({
            'success': True,
            'director_summary': stats.get('director_summary', {}),
            'top_stations': stats.get('stations', [])[:5]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

def open_browser():
    time.sleep(0.5)
    url = 'http://127.0.0.1:5050'
    try:
        if sys.platform == 'darwin':
            os.system(f'open "{url}"')
        elif sys.platform == 'win32':
            os.system(f'start "" "{url}"')
        else:
            webbrowser.open(url)
    except Exception:
        webbrowser.open(url)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    print("==========================================================")
    print(" Kiosk Hisobot Adminka Senior Web Dasturi ishga tushdi!")
    print(f" Manzil: http://127.0.0.1:{port}")
    print("==========================================================")
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', port=port, debug=False)
