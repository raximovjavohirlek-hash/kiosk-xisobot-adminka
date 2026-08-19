import openpyxl
import pandas as pd
from datetime import datetime, date

# 1. Exact Email to Station Mapping
EMAIL_MAP = {
    'andijonkiosk@railway.uz': ('Андижон', 4, 5),
    'buxorokiosk@railway.uz': ('Бухоро', 6, 7),
    'khivakiosk@railway.uz': ('Хива', 8, 9),
    'kiosk@axonlogic.uz': ('Тошкент Жанубий', 10, 11),
    'margilonkiosk@railway.uz': ('Марғилон', 12, 13),
    'navoiykiosk@railway.uz': ('Навои', 14, 15),
    'namangankiosk@railway.uz': ('Наманган', 16, 17),
    'nukuskiosk@railway.uz': ('Нукус', 18, 19),
    'qarshikiosk@railway.uz': ('Қарши', 20, 21),
    'qongirotkiosk@railway.uz': ('Қўнғирод', 22, 23),
    'qoqonkiosk@railway.uz': ('Қўқон', 24, 25),
    'samarqandkiosk@railway.uz': ('Самарқанд', 26, 27),
    'termizkiosk@railway.uz': ('Термиз', 28, 29),
    'toshkent.shimoliykiosk@railway.uz': ('Тошкент Марказий', 30, 31),
    'urganchkiosk@railway.uz': ('Урганч', 32, 33)
}

ONLINE_PAYMENTS = ['HamkorbankHold', 'HamkorbankWebView', 'Payme', 'StripeIntegration', 'OctoBankFC']
TERMINAL_PAYMENTS = ['Uzcard', 'Uzkassa', 'Kassa']

def generate_report():
    print("Reading transaction data from data.xlsx...")
    df = pd.read_excel('data.xlsx')
    
    df['Дата создания_dt'] = pd.to_datetime(df['Дата создания'], errors='coerce')
    df['Date'] = df['Дата создания_dt'].dt.date
    
    kiosk_df = df[df['Пользователь'].isin(EMAIL_MAP.keys())].copy()
    
    # Daily aggregation per station
    grouped_station = kiosk_df.groupby(['Date', 'Пользователь']).agg(
        tickets=('Количество билетов', 'sum'),
        summa=('Общая стоимость', 'sum')
    ).reset_index()
    
    # Daily payment method breakdown
    kiosk_df['PaymentType'] = kiosk_df['Способ оплаты'].apply(
        lambda x: 'Online' if x in ONLINE_PAYMENTS else 'Terminal'
    )
    
    grouped_payment = kiosk_df.groupby(['Date', 'PaymentType']).agg(
        tickets=('Количество билетов', 'sum'),
        summa=('Общая стоимость', 'sum')
    ).reset_index()
    
    wb = openpyxl.load_workbook('Август кисока.xlsx')
    
    ws_hudud = wb['Худудлар']
    ws_jami = wb['Жами']
    ws_list1 = wb['Лист1']
    ws_oylar = wb['Ойлар кесимида']
    
    # Map dates to row numbers in 'Худудлар' (Rows 4..34 for Aug 1..31)
    date_to_row = {}
    for r in range(4, 35):
        val = ws_hudud.cell(r, 1).value
        if isinstance(val, (datetime, date)):
            d_obj = val.date() if isinstance(val, datetime) else val
            date_to_row[d_obj] = r

    print("Updating 'Худудлар' and 'Жами' sheets...")
    for d in sorted(grouped_station['Date'].unique()):
        if d in date_to_row:
            r = date_to_row[d]
            day_df = grouped_station[grouped_station['Date'] == d]
            
            # Clear station cells for this day
            for c in range(4, 34):
                ws_hudud.cell(r, c).value = 0
                
            for email, (st_name, col_soni, col_summa) in EMAIL_MAP.items():
                match = day_df[day_df['Пользователь'] == email]
                if not match.empty:
                    t_val = int(match['tickets'].values[0])
                    s_val = int(match['summa'].values[0])
                    ws_hudud.cell(r, col_soni).value = t_val
                    ws_hudud.cell(r, col_summa).value = s_val

            # Ensure row formulas for total Soni (Col B) and total Summa (Col C)
            ws_hudud.cell(r, 2).value = f"=D{r}+F{r}+H{r}+J{r}+L{r}+N{r}+P{r}+R{r}+T{r}+V{r}+X{r}+Z{r}+AB{r}+AD{r}+AF{r}"
            ws_hudud.cell(r, 3).value = f"=E{r}+G{r}+I{r}+K{r}+M{r}+O{r}+Q{r}+S{r}+U{r}+W{r}+Y{r}+AA{r}+AC{r}+AE{r}+AG{r}"
            
            # Update 'Жами' sheet for this day
            ws_jami.cell(r, 2).value = f"=Худудлар!B{r}"
            ws_jami.cell(r, 3).value = f"=Худудлар!C{r}"
            
            # Payment breakdown
            day_pay = grouped_payment[grouped_payment['Date'] == d]
            on_row = day_pay[day_pay['PaymentType'] == 'Online']
            term_row = day_pay[day_pay['PaymentType'] == 'Terminal']
            
            on_tickets = int(on_row['tickets'].values[0]) if not on_row.empty else 0
            on_summa = int(on_row['summa'].values[0]) if not on_row.empty else 0
            term_tickets = int(term_row['tickets'].values[0]) if not term_row.empty else 0
            term_summa = int(term_row['summa'].values[0]) if not term_row.empty else 0
            
            ws_jami.cell(r, 4).value = on_tickets
            ws_jami.cell(r, 5).value = on_summa
            ws_jami.cell(r, 6).value = term_tickets
            ws_jami.cell(r, 7).value = term_summa

    # Ensure Row 35 SUM formulas in 'Худудлар'
    ws_hudud.cell(35, 2).value = "=SUM(B4:B34)"
    ws_hudud.cell(35, 3).value = "=SUM(C4:C34)"
    for c in range(4, 34):
        col_letter = openpyxl.utils.get_column_letter(c)
        ws_hudud.cell(35, c).value = f"=SUM({col_letter}4:{col_letter}34)"

    # Ensure Row 35 SUM formulas in 'Жами'
    ws_jami.cell(35, 2).value = "=SUM(B4:B34)"
    ws_jami.cell(35, 3).value = "=SUM(C4:C34)"
    ws_jami.cell(35, 4).value = "=SUM(D4:D34)"
    ws_jami.cell(35, 5).value = "=SUM(E4:E34)"
    ws_jami.cell(35, 6).value = "=SUM(F4:F34)"
    ws_jami.cell(35, 7).value = "=SUM(G4:G34)"

    # Calculate current totals for sorting Tables 1 & 2 in 'Лист1'
    station_sums = []
    for email, (st_name, col_soni, col_summa) in EMAIL_MAP.items():
        col_soni_letter = openpyxl.utils.get_column_letter(col_soni)
        col_summa_letter = openpyxl.utils.get_column_letter(col_summa)
        
        # Calculate actual sum from rows 4 to 34
        soni_val = 0
        summa_val = 0
        for r in range(4, 35):
            soni_val += (ws_hudud.cell(r, col_soni).value or 0)
            summa_val += (ws_hudud.cell(r, col_summa).value or 0)
            
        station_sums.append({
            'stansiya': st_name,
            'soni_val': soni_val,
            'summa_val': summa_val,
            'soni_formula': f"=Худудлар!{col_soni_letter}35",
            'summa_formula': f"=Худудлар!{col_summa_letter}35"
        })

    # Sort for Table 1 (by Summa descending)
    by_summa = sorted(station_sums, key=lambda x: x['summa_val'], reverse=True)
    for i, item in enumerate(by_summa, start=4):
        ws_list1.cell(i, 1).value = i - 3
        ws_list1.cell(i, 2).value = item['stansiya']
        ws_list1.cell(i, 3).value = item['soni_formula']
        ws_list1.cell(i, 4).value = item['summa_formula']
    ws_list1.cell(20, 3).value = "=SUM(C4:C18)"
    ws_list1.cell(20, 4).value = "=SUM(D4:D18)"

    # Sort for Table 2 (by Soni descending)
    by_soni = sorted(station_sums, key=lambda x: x['soni_val'], reverse=True)
    for i, item in enumerate(by_soni, start=27):
        ws_list1.cell(i, 1).value = i - 26
        ws_list1.cell(i, 2).value = item['stansiya']
        ws_list1.cell(i, 3).value = item['soni_formula']
        ws_list1.cell(i, 4).value = item['summa_formula']
    ws_list1.cell(43, 3).value = "=SUM(C27:C41)"
    ws_list1.cell(43, 4).value = "=SUM(D27:D41)"

    wb.save('Август кисока.xlsx')
    print("Report updated successfully!")

if __name__ == "__main__":
    generate_report()
