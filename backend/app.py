import re

def clean_int(val):
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str):
        cleaned = re.sub(r'[^\d\-]', '', val)
        if cleaned:
            try:
                return int(cleaned)
            except Exception:
                pass
    return 0

import os
import sys
import json
import time
import webbrowser
import threading
import io
import base64
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from datetime import datetime, date
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)  # Enable Cross-Origin Resource Sharing for Cloudflare Pages frontend
app.config['SECRET_KEY'] = 'kiosk-xisobot-secret-key-2026'
app.config['UPLOAD_FOLDER'] = os.path.dirname(os.path.abspath(__file__))
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB limit
app.config['ADMIN_PASSWORD'] = os.environ.get('ADMIN_PASSWORD', 'admin')

MAPPINGS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'kiosk_mappings.json')
UPLOAD_LOGS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'kiosk_upload_logs.json')
USERS_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'users.json')

def safe_copy_file(src, dst):
    if not src or not os.path.exists(src):
        return
    src_abs = os.path.abspath(src)
    dst_abs = os.path.abspath(dst)
    if src_abs == dst_abs:
        return
    os.makedirs(os.path.dirname(dst_abs), exist_ok=True)
    import shutil
    shutil.copy(src_abs, dst_abs)

def load_users():
    if os.path.exists(USERS_FILE):
        try:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    default_pass = app.config.get('ADMIN_PASSWORD', 'admin')
    return [{
        "username": "admin",
        "password": default_pass,
        "name": "Bosh Administrator",
        "role": "admin",
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }]

def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

DEFAULT_EMAIL_MAP = {
    "toshkent.shimoliykiosk@railway.uz": {"station": "Тошкент Марказий", "col_soni": 30, "col_summa": 31},
    "kiosk@axonlogic.uz": {"station": "Тошкент Жанубий", "col_soni": 10, "col_summa": 11},
    "samarqandkiosk@railway.uz": {"station": "Самарқанд", "col_soni": 26, "col_summa": 27},
    "urganchkiosk@railway.uz": {"station": "Урганч", "col_soni": 32, "col_summa": 33},
    "khivakiosk@railway.uz": {"station": "Хива", "col_soni": 8, "col_summa": 9},
    "navoiykiosk@railway.uz": {"station": "Навои", "col_soni": 14, "col_summa": 15},
    "buxorokiosk@railway.uz": {"station": "Бухоро", "col_soni": 6, "col_summa": 7},
    "qongirotkiosk@railway.uz": {"station": "Қўнғирод", "col_soni": 22, "col_summa": 23},
    "nukuskiosk@railway.uz": {"station": "Нукус", "col_soni": 18, "col_summa": 19},
    "andijonkiosk@railway.uz": {"station": "Андижон", "col_soni": 4, "col_summa": 5},
    "qoqonkiosk@railway.uz": {"station": "Қўқон", "col_soni": 24, "col_summa": 25},
    "margilonkiosk@railway.uz": {"station": "Марғилон", "col_soni": 12, "col_summa": 13},
    "namangankiosk@railway.uz": {"station": "Наманган", "col_soni": 16, "col_summa": 17},
    "termizkiosk@railway.uz": {"station": "Термиз", "col_soni": 28, "col_summa": 29},
    "qarshikiosk@railway.uz": {"station": "Қарши", "col_soni": 20, "col_summa": 21}
}

ONLINE_PAYMENTS = ['HamkorbankHold', 'HamkorbankWebView', 'Payme', 'StripeIntegration', 'OctoBankFC']

def load_mappings():
    if os.path.exists(MAPPINGS_FILE):
        try:
            with open(MAPPINGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return DEFAULT_EMAIL_MAP

def save_mappings(mappings):
    with open(MAPPINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(mappings, f, ensure_ascii=False, indent=2)

def load_upload_logs():
    if os.path.exists(UPLOAD_LOGS_FILE):
        try:
            with open(UPLOAD_LOGS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def add_upload_log(filename, rows_count, status="Muvaffaqiyatli"):
    logs = load_upload_logs()
    logs.insert(0, {
        "id": len(logs) + 1,
        "filename": filename,
        "rows": rows_count,
        "timestamp": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        "status": status
    })
    logs = logs[:50]
    with open(UPLOAD_LOGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

MONTH_NAMES_UZ = {
    1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel",
    5: "May", 6: "Iyun", 7: "Iyul", 8: "Avgust",
    9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"
}

def parse_report_file(fpath, fname):
    import zipfile, shutil
    if not os.path.exists(fpath) or not zipfile.is_zipfile(fpath):
        ex_path = os.path.join(app.config['UPLOAD_FOLDER'], 'excellar', fname)
        if os.path.exists(ex_path) and zipfile.is_zipfile(ex_path):
            try:
                shutil.copy(ex_path, fpath)
            except Exception:
                pass
    if not os.path.exists(fpath) or not zipfile.is_zipfile(fpath):
        return None, None
    email_map = load_mappings()
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        if 'Худудлар' not in wb.sheetnames:
            return None, None
        ws_h = wb['Худудлар']
        ws_j = wb['Жами'] if 'Жами' in wb.sheetnames else None

        fn_lower = fname.lower()
        if 'январ' in fn_lower: ym_code = '2026-01'
        elif 'феврал' in fn_lower: ym_code = '2026-02'
        elif 'март' in fn_lower: ym_code = '2026-03'
        elif 'апрел' in fn_lower: ym_code = '2026-04'
        elif 'маи' in fn_lower or 'май' in fn_lower: ym_code = '2026-05'
        elif 'июн' in fn_lower: ym_code = '2026-06'
        elif 'июл' in fn_lower: ym_code = '2026-07'
        elif 'август' in fn_lower: ym_code = '2026-08'
        else:
            return None, None

        import re
        STATION_SYNONYMS = {
            'хива': 'Хива', 'самарканд': 'Самарқанд', 'самарқанд': 'Самарқанд',
            'тошкент марказий': 'Тошкент Марказий', 'тошкент марказий ': 'Тошкент Марказий',
            'тошкент жанубий': 'Тошкент Жанубий', 'навои': 'Навои', 'навоий': 'Навои',
            'бухоро': 'Бухоро', 'нукус': 'Нукус', 'урганч': 'Урганч', 'ургенч': 'Урганч',
            'карши': 'Қарши', 'қарши': 'Қарши', 'термез': 'Термиз', 'термиз': 'Термиз',
            'кунгрод': 'Қўнғирод', 'қўнғирод': 'Қўнғирод', 'андижон': 'Андижон',
            'қўқон': 'Қўқон', 'кокон': 'Қўқон', 'марғилон': 'Марғилон', 'маргилон': 'Марғилон',
            'наманган': 'Наманган'
        }

        # Dynamically map station columns from sheet headers (Row 2 / Row 1)
        dynamic_cols = {}
        for c in range(2, ws_h.max_column + 1):
            v1 = ws_h.cell(1, c).value
            v2 = ws_h.cell(2, c).value
            for raw_v in [v2, v1]:
                if raw_v:
                    clean_v = str(raw_v).strip().lower()
                    clean_v = re.sub(r'\s+', ' ', clean_v)
                    normalized = STATION_SYNONYMS.get(clean_v)
                    if normalized and normalized not in dynamic_cols:
                        dynamic_cols[normalized] = (c, c + 1)
                        break

        station_totals = {email: {'tickets': 0, 'summa': 0, 'daily': []} for email in email_map}
        daily_trend = []

        for r in range(4, 35):
            row_date = ws_h.cell(r, 1).value
            if isinstance(row_date, str):
                s_val = row_date.strip().lower()
                if 'жам' in s_val or 'jam' in s_val or 'total' in s_val:
                    break
            if not row_date:
                m_num = ym_code.split('-')[1]
                y_num = ym_code.split('-')[0]
                day_num = r - 3
                d_str = f"{day_num:02d}.{m_num}.{y_num}"
            else:
                d_str = row_date.strftime('%d.%m.%Y') if hasattr(row_date, 'strftime') else str(row_date)

            day_tickets = 0
            day_summa = 0
            for email, meta in email_map.items():
                st_name = meta['station']
                if st_name in dynamic_cols:
                    c_soni, c_summa = dynamic_cols[st_name]
                else:
                    c_soni = meta['col_soni']
                    c_summa = meta['col_summa']

                v_soni = clean_int(ws_h.cell(r, c_soni).value)
                v_summa = clean_int(ws_h.cell(r, c_summa).value)

                station_totals[email]['tickets'] += v_soni
                station_totals[email]['summa'] += v_summa
                station_totals[email]['daily'].append({
                    'date': d_str,
                    'tickets': v_soni,
                    'summa': v_summa
                })
                day_tickets += v_soni
                day_summa += v_summa

            raw_on_t = ws_j.cell(r, 4).value if ws_j else 0
            raw_on_s = ws_j.cell(r, 5).value if ws_j else 0
            raw_term_t = ws_j.cell(r, 6).value if ws_j else 0
            raw_term_s = ws_j.cell(r, 7).value if ws_j else 0

            term_t = clean_int(raw_term_t)
            term_s = clean_int(raw_term_s)
            on_t = clean_int(raw_on_t) if raw_on_t else max(0, day_tickets - term_t)
            on_s = clean_int(raw_on_s) if raw_on_s else max(0, day_summa - term_s)

            daily_trend.append({
                'date': d_str, 'tickets': day_tickets, 'summa': day_summa,
                'online_tickets': on_t, 'online_summa': on_s,
                'terminal_tickets': term_t, 'terminal_summa': term_s
            })

        station_sums = []
        m_total_tickets = sum(station_totals[email]['tickets'] for email in email_map)
        m_total_summa = sum(station_totals[email]['summa'] for email in email_map)

        for email, meta in email_map.items():
            st_name = meta['station']
            s_val = station_totals[email]['summa']
            t_val = station_totals[email]['tickets']
            sh_pct = round((s_val / m_total_summa) * 100, 1) if m_total_summa > 0 else 0
            station_sums.append({
                'stansiya': st_name, 'email': email,
                'soni_val': t_val,
                'summa_val': s_val,
                'share_percent': sh_pct,
                'daily_breakdown': station_totals[email]['daily']
            })

        by_summa = sorted(station_sums, key=lambda x: x['summa_val'], reverse=True)

        return ym_code, {
            'total_tickets': m_total_tickets, 'total_summa': m_total_summa,
            'stations': by_summa, 'daily_trend': daily_trend
        }
    except Exception as ex:
        ex_path = os.path.join(app.config['UPLOAD_FOLDER'], 'excellar', fname)
        if fpath != ex_path and os.path.exists(ex_path) and zipfile.is_zipfile(ex_path):
            try:
                shutil.copy(ex_path, fpath)
                return parse_report_file(fpath, fname)
            except Exception:
                pass
        print("parse_report_file error:", fname, ex)
        return None, None

def get_all_official_monthly_reports():
    reports = {}
    search_paths = [
        app.config['UPLOAD_FOLDER'],
        os.path.join(app.config['UPLOAD_FOLDER'], 'excellar')
    ]
    for sp in search_paths:
        if os.path.exists(sp):
            for fn in sorted(os.listdir(sp)):
                if fn.endswith('.xlsx') and not fn.startswith('~$'):
                    fn_lower = fn.lower()
                    if fn_lower.startswith(('data', 'orders', 'export')) and ('киоска' not in fn_lower and 'кисока' not in fn_lower):
                        continue
                    fp = os.path.join(sp, fn)
                    ym, stats = parse_report_file(fp, fn)
                    if ym and stats and ym not in reports:
                        reports[ym] = stats
    return reports

STATS_CACHE = None

def invalidate_stats_cache():
    global STATS_CACHE
    STATS_CACHE = None

def process_excel(data_path, report_path, uploaded_path=None):
    email_map = load_mappings()
    
    # 1. Load official monthly excel reports from backend/excellar
    monthly_data = get_all_official_monthly_reports()

    # 2. Try parsing uploaded files as official monthly report files
    raw_candidates = [uploaded_path, data_path, report_path]
    candidate_paths = []
    for cp in raw_candidates:
        if cp and os.path.exists(cp):
            abs_p = os.path.abspath(cp)
            if abs_p not in candidate_paths:
                candidate_paths.append(abs_p)

    for cp in candidate_paths:
        fname = os.path.basename(cp)
        fn_lower = fname.lower()
        if fn_lower.startswith(('data', 'orders', 'export')) and ('киоска' not in fn_lower and 'кисока' not in fn_lower):
            continue
        ym, rep_stats = parse_report_file(cp, fname)
        if ym and rep_stats:
            existing_tix = monthly_data.get(ym, {}).get('total_tickets', 0)
            if ym not in monthly_data or rep_stats.get('total_tickets', 0) >= existing_tix:
                monthly_data[ym] = rep_stats

    # 3. Try parsing uploaded files as raw transaction data
    for cp in candidate_paths:
        if cp and os.path.exists(cp):
            try:
                df = pd.read_excel(cp)
                if 'Дата создания' in df.columns:
                    df['Дата создания_dt'] = pd.to_datetime(df['Дата создания'], errors='coerce')
                    df['Date'] = df['Дата создания_dt'].dt.date
                    df['YearMonth'] = df['Дата создания_dt'].dt.strftime('%Y-%m')
                    
                    kiosk_df = df[df['Пользователь'].isin(email_map.keys())].copy()
                    kiosk_df['PaymentType'] = kiosk_df['Способ оплаты'].apply(
                        lambda x: 'Online' if x in ONLINE_PAYMENTS else 'Terminal'
                    )
                    
                    unique_periods = sorted([p for p in kiosk_df['YearMonth'].dropna().unique()], reverse=True)
                    for ym in unique_periods:
                        m_kiosk_df = kiosk_df[kiosk_df['YearMonth'] == ym]
                        
                        m_grouped_station = m_kiosk_df.groupby(['Date', 'Пользователь']).agg(
                            tickets=('Количество билетов', 'sum'),
                            summa=('Общая стоимость', 'sum')
                        ).reset_index()
                        
                        m_grouped_payment = m_kiosk_df.groupby(['Date', 'PaymentType']).agg(
                            tickets=('Количество билетов', 'sum'),
                            summa=('Общая стоимость', 'sum')
                        ).reset_index()
                        
                        m_station_sums = []
                        for email, meta in email_map.items():
                            st_name = meta['station']
                            m_match = m_grouped_station[m_grouped_station['Пользователь'] == email]
                            soni_val = int(m_match['tickets'].sum()) if not m_match.empty else 0
                            summa_val = int(m_match['summa'].sum()) if not m_match.empty else 0
                            m_station_sums.append({
                                'stansiya': st_name,
                                'email': email,
                                'soni_val': soni_val,
                                'summa_val': summa_val
                            })
                            
                        m_by_summa = sorted(m_station_sums, key=lambda x: x['summa_val'], reverse=True)
                        m_total_tickets = sum(s['soni_val'] for s in m_station_sums)
                        m_total_summa = sum(s['summa_val'] for s in m_station_sums)
                        
                        m_daily_trend = []
                        all_dates_in_m = sorted(m_kiosk_df['Date'].dropna().unique())
                        for d in all_dates_in_m:
                            d_str = d.strftime('%d.%m.%Y')
                            d_st_df = m_grouped_station[m_grouped_station['Date'] == d]
                            d_tickets = int(d_st_df['tickets'].sum())
                            d_summa = int(d_st_df['summa'].sum())
                            
                            d_pay_df = m_grouped_payment[m_grouped_payment['Date'] == d]
                            on_row = d_pay_df[d_pay_df['PaymentType'] == 'Online']
                            term_row = d_pay_df[d_pay_df['PaymentType'] == 'Terminal']
                            
                            on_t = int(on_row['tickets'].values[0]) if not on_row.empty else 0
                            on_s = int(on_row['summa'].values[0]) if not on_row.empty else 0
                            term_t = int(term_row['tickets'].values[0]) if not term_row.empty else 0
                            term_s = int(term_row['summa'].values[0]) if not term_row.empty else 0
                            
                            m_daily_trend.append({
                                'date': d_str,
                                'tickets': d_tickets,
                                'summa': d_summa,
                                'online_tickets': on_t,
                                'online_summa': on_s,
                                'terminal_tickets': term_t,
                                'terminal_summa': term_s
                            })
                            
                        existing_tix = monthly_data.get(ym, {}).get('total_tickets', 0)
                        if ym not in monthly_data or m_total_tickets >= existing_tix:
                            monthly_data[ym] = {
                                'total_tickets': m_total_tickets,
                                'total_summa': m_total_summa,
                                'stations': m_by_summa,
                                'daily_trend': m_daily_trend
                            }
            except Exception as ex:
                print("data_path parse warning:", ex)

    # Build available_months sorted descending
    all_ym_codes = sorted(list(monthly_data.keys()), reverse=True)
    available_months = []
    for ym in all_ym_codes:
        try:
            y, m = ym.split('-')
            m_name = MONTH_NAMES_UZ.get(int(m), ym)
            available_months.append({
                'code': ym,
                'name': f"{m_name} {y}"
            })
        except Exception:
            available_months.append({'code': ym, 'name': ym})

    # All-time / overall totals across all months in monthly_data
    overall_station_totals = {email: {'tickets': 0, 'summa': 0, 'daily': []} for email in email_map}
    overall_daily_trend = []
    
    # YTD totals (latest year)
    latest_year = all_ym_codes[0].split('-')[0] if all_ym_codes else '2026'
    ytd_station_totals = {email: {'tickets': 0, 'summa': 0, 'daily': []} for email in email_map}
    ytd_daily_trend = []

    for ym, m_info in monthly_data.items():
        for st in (m_info.get('stations') or []):
            em = st.get('email')
            if em in overall_station_totals:
                overall_station_totals[em]['tickets'] += st.get('soni_val', 0)
                overall_station_totals[em]['summa'] += st.get('summa_val', 0)
                if st.get('daily_breakdown'):
                    overall_station_totals[em]['daily'].extend(st.get('daily_breakdown'))
                if ym.startswith(latest_year):
                    ytd_station_totals[em]['tickets'] += st.get('soni_val', 0)
                    ytd_station_totals[em]['summa'] += st.get('summa_val', 0)
                    if st.get('daily_breakdown'):
                        ytd_station_totals[em]['daily'].extend(st.get('daily_breakdown'))
        overall_daily_trend.extend(m_info.get('daily_trend') or [])
        if ym.startswith(latest_year):
            ytd_daily_trend.extend(m_info.get('daily_trend') or [])

    all_station_sums = []
    ytd_station_sums = []
    for email, meta in email_map.items():
        st_name = meta['station']
        all_station_sums.append({
            'stansiya': st_name,
            'email': email,
            'soni_val': overall_station_totals[email]['tickets'],
            'summa_val': overall_station_totals[email]['summa'],
            'daily_breakdown': overall_station_totals[email]['daily']
        })
        ytd_station_sums.append({
            'stansiya': st_name,
            'email': email,
            'soni_val': ytd_station_totals[email]['tickets'],
            'summa_val': ytd_station_totals[email]['summa'],
            'daily_breakdown': ytd_station_totals[email]['daily']
        })

    overall_by_summa = sorted(all_station_sums, key=lambda x: x['summa_val'], reverse=True)
    overall_total_tickets = sum(s['soni_val'] for s in all_station_sums)
    overall_total_summa = sum(s['summa_val'] for s in all_station_sums)

    overall_data_map = {
        'total_tickets': overall_total_tickets,
        'total_summa': overall_total_summa,
        'stations': overall_by_summa,
        'daily_trend': overall_daily_trend
    }

    ytd_by_summa = sorted(ytd_station_sums, key=lambda x: x['summa_val'], reverse=True)
    ytd_total_tickets = sum(s['soni_val'] for s in ytd_station_sums)
    ytd_total_summa = sum(s['summa_val'] for s in ytd_station_sums)

    ytd_data_map = {
        'total_tickets': ytd_total_tickets,
        'total_summa': ytd_total_summa,
        'stations': ytd_by_summa,
        'daily_trend': ytd_daily_trend,
        'year': latest_year
    }

    return enrich_stats_with_executive_metrics(monthly_data, overall_data_map, ytd_data_map, available_months)

def enrich_stats_with_executive_metrics(monthly_data_map, overall_data_map, ytd_data_map, available_months):
    def format_stations(station_list, total_summa):
        enriched = []
        for st in station_list:
            soni = st.get('soni_val', 0)
            summa = st.get('summa_val', 0)
            avg_price = round(summa / soni) if soni > 0 else 0
            share_pct = round((summa / total_summa * 100), 1) if total_summa > 0 else 0.0
            
            st_copy = dict(st)
            st_copy.update({
                'avg_price': avg_price,
                'share_percent': share_pct
            })
            enriched.append(st_copy)
        return enriched

    def build_summary(stats_dict, period_name="ushbu davr"):
        t_sum = stats_dict.get('total_summa', 0)
        t_tix = stats_dict.get('total_tickets', 0)
        avg_p = round(t_sum / t_tix) if t_tix > 0 else 0

        d_trend = stats_dict.get('daily_trend', [])
        d_len = len(d_trend)
        d_avg_s = round(t_sum / d_len) if d_len > 0 else 0
        d_avg_t = round(t_tix / d_len) if d_len > 0 else 0

        peak_day = max(d_trend, key=lambda x: x.get('summa', 0)) if d_trend else {'date': '-', 'summa': 0, 'tickets': 0}

        st_list = stats_dict.get('stations', [])
        top_st = st_list[0] if st_list else {'stansiya': "Noma'lum", 'summa_val': 0, 'soni_val': 0, 'share_percent': 0}
        sec_st = st_list[1] if len(st_list) > 1 else {'stansiya': "-", 'summa_val': 0, 'share_percent': 0}

        on_t = sum(d.get('online_tickets', 0) for d in d_trend)
        term_t = sum(d.get('terminal_tickets', 0) for d in d_trend)
        tot_pay = on_t + term_t or 1
        on_p = round((on_t / tot_pay) * 100, 1)

        return {
            'net_revenue': t_sum,
            'total_tickets': t_tix,
            'overall_avg_price': avg_p,
            'daily_avg_revenue': d_avg_s,
            'daily_avg_tickets': d_avg_t,
            'peak_date': peak_day.get('date', '-'),
            'peak_day_revenue': peak_day.get('summa', 0),
            'top_station': top_st.get('stansiya'),
            'top_station_summa': top_st.get('summa_val', 0),
            'top_station_share': top_st.get('share_percent', 0),
            'second_station': sec_st.get('stansiya'),
            'second_station_summa': sec_st.get('summa_val', 0),
            'online_percent': on_p,
            'terminal_percent': round(100 - on_p, 1) if on_p else 0.0,
            'period_name': period_name,
            'ai_recommendation': f"Hurmatli Rahbariyat, {period_name} bo'yicha kiosklar orqali jami {t_sum:,} so'm tushum hamda {t_tix:,} ta chipta sotildi. "
                                f"Bitta chiptaning o'rtacha narxi {avg_p:,} so'mni va kunlik o'rtacha tushum {d_avg_s:,} so'mni tashkil etdi. "
                                f"Eng savdoli kassa {top_st.get('stansiya')} bo'lib, uning umumiy tushumdagi ulushi {top_st.get('share_percent')}% ni tashkil qiladi. "
                                f"Eng yuqori kunlik savdo ko'rsatkichi {peak_day.get('date')} sanasida ({peak_day.get('summa', 0):,} so'm) qayd etilgan."
        }

    for ym, m_info in monthly_data_map.items():
        t_sum = m_info.get('total_summa', 0)
        m_info['stations'] = format_stations(m_info.get('stations', []), t_sum)
        month_name = ym
        for m_obj in available_months:
            if m_obj['code'] == ym:
                month_name = m_obj['name']
                break
        m_info['director_summary'] = build_summary(m_info, f"{month_name} oyi")

    tot_sum = overall_data_map.get('total_summa', 0)
    overall_data_map['stations'] = format_stations(overall_data_map.get('stations', []), tot_sum)
    overall_data_map['director_summary'] = build_summary(overall_data_map, "barcha oylar birgalikda (Jami)")

    ytd_tot_sum = ytd_data_map.get('total_summa', 0)
    ytd_data_map['stations'] = format_stations(ytd_data_map.get('stations', []), ytd_tot_sum)
    ytd_year = ytd_data_map.get('year', '2026')
    ytd_data_map['director_summary'] = build_summary(ytd_data_map, f"{ytd_year}-yil boshidan beri (YTD)")

    latest_ym = available_months[0]['code'] if available_months else None
    default_stats = monthly_data_map.get(latest_ym) if latest_ym else overall_data_map

    executive_summary = default_stats.get('director_summary') or build_summary(default_stats, "hozirgi oy")

    return {
        'total_tickets': default_stats['total_tickets'],
        'total_summa': default_stats['total_summa'],
        'stations': default_stats['stations'],
        'daily_trend': default_stats['daily_trend'],
        'available_months': available_months,
        'monthly_data': monthly_data_map,
        'overall_data': overall_data_map,
        'ytd_data': ytd_data_map,
        'director_summary': executive_summary,
        'last_updated': datetime.now().strftime('%d.%m.%Y %H:%M')
    }

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api-sync')
@app.route('/admin/api-sync')
def api_sync_page():
    return render_template('api_sync.html')

@app.route('/api/stats', methods=['GET'])
def get_stats():
    global STATS_CACHE
    if STATS_CACHE is not None:
        return jsonify({'success': True, 'stats': STATS_CACHE, 'monthly_reports': STATS_CACHE.get('monthly_data', {})})

    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
    try:
        stats = process_excel(data_path, report_path, uploaded_path=save_path)
        STATS_CACHE = stats
        return jsonify({'success': True, 'stats': stats, 'monthly_reports': stats.get('monthly_data', {})})
    except Exception as e:
        print("get_stats error:", e)
        return jsonify({'success': False, 'error': str(e)}), 500

def safe_filename(filename):
    filename = os.path.basename(filename)
    clean_name = "".join(c for c in filename if c.isprintable() and c not in '/\\:*?"<>|')
    return clean_name or "uploaded_file.xlsx"

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'Fayl tanlanmagan'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'Fayl tanlanmagan'}), 400

    orig_filename = file.filename
    filename = safe_filename(orig_filename)
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        file.save(save_path)

        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
        data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')

        import zipfile
        if not os.path.exists(report_path) or not zipfile.is_zipfile(report_path):
            ex_backup = os.path.join(app.config['UPLOAD_FOLDER'], 'excellar', 'Август кисока.xlsx')
            if os.path.exists(ex_backup) and zipfile.is_zipfile(ex_backup):
                safe_copy_file(ex_backup, report_path)

        fn_lower = filename.lower()
        orig_lower = orig_filename.lower()
        rows_count = 0

        if 'кисока' in orig_lower or 'киоска' in orig_lower or 'кисока' in fn_lower or 'киоска' in fn_lower:
            ex_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'excellar')
            os.makedirs(ex_dir, exist_ok=True)
            safe_copy_file(save_path, os.path.join(ex_dir, filename))
            safe_copy_file(save_path, report_path)
            rows_count = 1
        else:
            try:
                df = pd.read_excel(save_path)
                rows_count = len(df)
                safe_copy_file(save_path, data_path)
                backend_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'backend')
                if os.path.exists(backend_dir):
                    safe_copy_file(save_path, os.path.join(backend_dir, 'data.xlsx'))
            except Exception as parse_ex:
                print("Uploaded raw data parse warning:", parse_ex)
                safe_copy_file(save_path, data_path)
                if os.path.exists(data_path):
                    try:
                        rows_count = len(pd.read_excel(data_path))
                    except Exception:
                        rows_count = 1

        invalidate_stats_cache()
        stats = process_excel(data_path, report_path, uploaded_path=save_path)
        global STATS_CACHE
        STATS_CACHE = stats
        
        add_upload_log(orig_filename, rows_count if rows_count > 0 else 1, "Muvaffaqiyatli")
        
        return jsonify({
            'success': True, 
            'message': f"Fayl muvaffaqiyatli yuklandi va ma'lumotlar yangilandi! ({rows_count} yozuv)", 
            'stats': stats
        })
    except Exception as e:
        add_upload_log(orig_filename, 0, f"Xatolik: {str(e)}")
        return jsonify({'success': False, 'error': f"Xatolik yuz berdi: {str(e)}"}), 500

@app.route('/api/download', methods=['GET'])
def download():
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    if os.path.exists(report_path):
        return send_file(report_path, as_attachment=True, download_name='Август_киоска_hisobot.xlsx')
    return jsonify({'error': 'Fayl topilmadi'}), 404

@app.route('/api/export-station-excel/<path:station_name>', methods=['GET'])
def export_station_excel(station_name):
    try:
        requested_month = request.args.get('month')
        all_reports = get_all_official_monthly_reports()
        
        if requested_month and requested_month in all_reports:
            stats = all_reports[requested_month]
            m_code_str = requested_month
        else:
            report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
            data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
            stats = process_excel(data_path, report_path, uploaded_path=save_path)
            m_code_str = '2026-08'
        
        stations = stats.get('stations', [])
        station_data = None
        rank = 0
        for idx, s in enumerate(stations):
            if s.get('stansiya') == station_name:
                station_data = s
                rank = idx + 1
                break
        
        if not station_data:
            return jsonify({'error': 'Stansiya topilmadi'}), 404

        # Create openpyxl workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Hisobot - {station_name[:20]}"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name='Calibri', size=15, bold=True, color='1F2937')
        subtitle_font = Font(name='Calibri', size=11, italic=True, color='4B5563')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Deep Blue
        
        kpi_title_font = Font(name='Calibri', size=10, bold=True, color='4B5563')
        kpi_val_font = Font(name='Calibri', size=13, bold=True, color='1E3A8A')
        kpi_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        
        total_font = Font(name='Calibri', size=11, bold=True, color='065F46')
        total_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid') # Emerald
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_right = Alignment(horizontal='right', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        # Title Block
        ws.merge_cells('A1:E1')
        ws['A1'] = f"O'ZBEKISTON TEMIR YO'LLARI — KIOSK HISOBOTI"
        ws['A1'].font = title_font
        ws['A1'].alignment = align_left

        ws.merge_cells('A2:E2')
        ws['A2'] = f"Kassa / Stansiya: {station_data['stansiya']} ({rank}-O'rin) | Shakllangan vaqt: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = align_left

        # KPI Cards Block (Rows 4-5)
        ws['A4'] = "Jami Tushum Summasi"
        ws['A5'] = f"{station_data['summa_val']:,} so'm"
        
        ws['B4'] = "Sotilgan Chiptalar"
        ws['B5'] = f"{station_data['soni_val']:,} ta"
        
        avg_p = round(station_data['summa_val'] / station_data['soni_val']) if station_data['soni_val'] > 0 else 0
        ws['C4'] = "O'rtacha Chek"
        ws['C5'] = f"{avg_p:,} so'm"
        
        ws['D4'] = "Tushum Ulushi"
        ws['D5'] = f"{station_data.get('share_percent', 0)}%"

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}4'].font = kpi_title_font
            ws[f'{col}4'].fill = kpi_fill
            ws[f'{col}4'].alignment = align_center
            ws[f'{col}5'].font = kpi_val_font
            ws[f'{col}5'].fill = kpi_fill
            ws[f'{col}5'].alignment = align_center

        # Table Headers (Row 7)
        headers = ["№", "Sana", "Sotilgan Chiptalar (ta)", "Kunlik Tushum (so'm)", "O'rtacha Chek (so'm)"]
        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        # Table Data (Row 8+)
        daily_list = station_data.get('daily_breakdown', [])
        row_idx = 8
        total_tickets = 0
        total_sum = 0

        for i, day in enumerate(daily_list, 1):
            t_val = day.get('tickets', 0)
            s_val = day.get('summa', 0)
            day_avg = round(s_val / t_val) if t_val > 0 else 0
            
            total_tickets += t_val
            total_sum += s_val

            ws.cell(row=row_idx, column=1, value=i).alignment = align_center
            ws.cell(row=row_idx, column=2, value=day.get('date', '')).alignment = align_center
            
            c3 = ws.cell(row=row_idx, column=3, value=t_val)
            c3.alignment = align_right
            c3.number_format = '#,##0'
            
            c4 = ws.cell(row=row_idx, column=4, value=s_val)
            c4.alignment = align_right
            c4.number_format = '#,##0" so\'m"'
            
            c5 = ws.cell(row=row_idx, column=5, value=day_avg)
            c5.alignment = align_right
            c5.number_format = '#,##0" so\'m"'

            for c in range(1, 6):
                ws.cell(row=row_idx, column=c).border = thin_border

            row_idx += 1

        # Total Row
        grand_avg = round(total_sum / total_tickets) if total_tickets > 0 else 0
        ws.cell(row=row_idx, column=1, value="—").alignment = align_center
        
        c2 = ws.cell(row=row_idx, column=2, value="JAMI (YAKUNIY)")
        c2.font = total_font
        c2.alignment = align_left
        
        c3 = ws.cell(row=row_idx, column=3, value=total_tickets)
        c3.font = total_font
        c3.alignment = align_right
        c3.number_format = '#,##0'

        c4 = ws.cell(row=row_idx, column=4, value=total_sum)
        c4.font = total_font
        c4.alignment = align_right
        c4.number_format = '#,##0" so\'m"'

        c5 = ws.cell(row=row_idx, column=5, value=grand_avg)
        c5.font = total_font
        c5.alignment = align_right
        c5.number_format = '#,##0" so\'m"'

        for c in range(1, 6):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = total_fill
            cell.border = thin_border

        # Adjust Column Widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 24
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        safe_st_name = "".join(c for c in station_name if c.isalnum() or c in (' ', '_', '-')).strip()
        download_name = f"Kiosk_Hisobot_{safe_st_name}_Avgust_2026.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print("export_station_excel error:", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/mappings', methods=['GET', 'POST'])
def handle_mappings():
    if request.method == 'POST':
        new_map = request.json
        save_mappings(new_map)
        invalidate_stats_cache()
        return jsonify({'success': True, 'message': 'Pochta biriktirmalari saqlandi!'})
    return jsonify({'success': True, 'mappings': load_mappings()})

@app.route('/api/upload-logs', methods=['GET'])
def get_upload_logs():
    return jsonify({'success': True, 'logs': load_upload_logs()})

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json or {}
    password = data.get('password', '')
    if password == app.config['ADMIN_PASSWORD']:
        return jsonify({'success': True, 'message': 'Admin rejimiga kirdingiz!'})
    return jsonify({'success': False, 'error': "Parol noto'g'ri!"}), 401

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.json or {}
    username = str(data.get('username', '')).strip().lower()
    password = str(data.get('password', '')).strip()
    
    users = load_users()
    for u in users:
        if str(u.get('username', '')).strip().lower() == username and str(u.get('password', '')).strip() == password:
            token = base64.b64encode(f"{username}:{time.time()}".encode()).decode()
            return jsonify({
                'success': True,
                'message': 'Muvaffaqiyatli tizimga kirdingiz!',
                'token': token,
                'user': {
                    'username': u.get('username'),
                    'name': u.get('name', u.get('username')),
                    'role': u.get('role', 'user')
                }
            })
    
    if (username == 'admin' or not username) and password == app.config['ADMIN_PASSWORD']:
        token = base64.b64encode(f"admin:{time.time()}".encode()).decode()
        return jsonify({
            'success': True,
            'message': 'Bosh administrator sifatida kirdingiz!',
            'token': token,
            'user': {
                'username': 'admin',
                'name': 'Bosh Administrator',
                'role': 'admin'
            }
        })
        
    return jsonify({'success': False, 'error': "Login yoki parol noto'g'ri!"}), 401

@app.route('/api/users', methods=['GET', 'POST'])
def manage_users():
    if request.method == 'GET':
        users = load_users()
        safe_users = [{
            'username': u.get('username'),
            'name': u.get('name'),
            'role': u.get('role', 'user'),
            'created_at': u.get('created_at', '')
        } for u in users]
        return jsonify({'success': True, 'users': safe_users})
    
    elif request.method == 'POST':
        data = request.json or {}
        username = str(data.get('username', '')).strip().lower()
        password = str(data.get('password', '')).strip()
        name = str(data.get('name', '')).strip() or username
        role = str(data.get('role', 'user')).strip().lower()
        
        if not username or not password:
            return jsonify({'success': False, 'error': "Login va parol kiritilishi shart!"}), 400
            
        users = load_users()
        if any(u.get('username', '').lower() == username for u in users):
            return jsonify({'success': False, 'error': "Bunday loginli foydalanuvchi allaqachon mavjud!"}), 400
            
        users.append({
            'username': username,
            'password': password,
            'name': name,
            'role': role,
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        save_users(users)
        return jsonify({'success': True, 'message': f"Foydalanuvchi '{username}' muvaffaqiyatli qo'shildi!"})

@app.route('/api/users/<username>', methods=['DELETE'])
def delete_user(username):
    username_clean = str(username).strip().lower()
    if username_clean == 'admin':
        return jsonify({'success': False, 'error': "Bosh admin foydalanuvchisini o'chirib bo'lmaydi!"}), 400
        
    users = load_users()
    new_users = [u for u in users if u.get('username', '').lower() != username_clean]
    if len(new_users) == len(users):
        return jsonify({'success': False, 'error': "Foydalanuvchi topilmadi!"}), 404
        
    save_users(new_users)
    return jsonify({'success': True, 'message': f"Foydalanuvchi '{username_clean}' o'chirildi!"})

TOKEN_FILE = os.path.join(app.config['UPLOAD_FOLDER'], 'kiosk_token.json')

def load_stored_token():
    if os.path.exists(TOKEN_FILE):
        try:
            with open(TOKEN_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {
                    'token': data.get('token', ''),
                    'csrf_token': data.get('csrf_token', '86de5ba7-734b-4b43-93a9-6ebcd5bfaa40'),
                    'cookie': data.get('cookie', '')
                }
        except Exception:
            pass
    return {
        'token': '',
        'csrf_token': '86de5ba7-734b-4b43-93a9-6ebcd5bfaa40',
        'cookie': ''
    }

def save_stored_token(token, csrf_token=None, cookie=None):
    current = load_stored_token()
    new_token = token if token is not None else current.get('token', '')
    new_csrf = csrf_token if csrf_token is not None else current.get('csrf_token', '86de5ba7-734b-4b43-93a9-6ebcd5bfaa40')
    new_cookie = cookie if cookie is not None else current.get('cookie', '')
    with open(TOKEN_FILE, 'w', encoding='utf-8') as f:
        json.dump({
            'token': new_token,
            'csrf_token': new_csrf,
            'cookie': new_cookie,
            'updated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }, f, ensure_ascii=False, indent=2)

def check_jwt_health(token, csrf_token=None):
    if not token or not token.strip():
        return {
            'valid': False,
            'status': 'missing',
            'message': "Bearer Token kiritilmagan",
            'expires_in_minutes': 0,
            'exp_datetime': None
        }
    return {
        'valid': True,
        'status': 'active',
        'message': "Token va CSRF kiritilgan (Faol)",
        'expires_in_minutes': 60,
        'exp_datetime': None
    }

@app.route('/api/admin/token', methods=['GET', 'POST'])
def manage_token():
    if request.method == 'POST':
        data = request.json or {}
        token = data.get('token', '').strip()
        csrf_token = data.get('csrf_token', '').strip()
        cookie = data.get('cookie', '').strip()
        save_stored_token(token, csrf_token, cookie)
        health = check_jwt_health(token, csrf_token)
        return jsonify({
            'success': True,
            'message': 'Bearer Token va Sozlamalar saqlandi!',
            'health': health
        })
    else:
        info = load_stored_token()
        token = info.get('token', '')
        csrf_token = info.get('csrf_token', '')
        cookie = info.get('cookie', '')
        health = check_jwt_health(token, csrf_token)
        masked_token = (token[:15] + '...' + token[-10:]) if len(token) > 30 else token
        return jsonify({
            'success': True,
            'token': token,
            'csrf_token': csrf_token,
            'cookie': cookie,
            'masked_token': masked_token,
            'health': health
        })

@app.route('/api/admin/token-health', methods=['GET'])
def get_token_health():
    info = load_stored_token()
    token = info.get('token', '')
    csrf_token = info.get('csrf_token', '')
    health = check_jwt_health(token, csrf_token)
    return jsonify({
        'success': True,
        'health': health
    })

@app.route('/api/admin/fetch-api-excel', methods=['POST'])
def fetch_api_excel():
    data = request.json or {}
    start_date = data.get('startDate', '').strip()
    end_date = data.get('endDate', '').strip()
    custom_token = data.get('token', '').strip()
    custom_csrf = data.get('csrf_token', '').strip()
    custom_cookie = data.get('cookie', '').strip()
    
    info = load_stored_token()
    token = custom_token if custom_token else info.get('token', '')
    csrf_token = custom_csrf if custom_csrf else info.get('csrf_token', '86de5ba7-734b-4b43-93a9-6ebcd5bfaa40')
    cookie_str = custom_cookie if custom_cookie else info.get('cookie', '')

    if not token or not token.strip():
        return jsonify({
            'success': False,
            'error': "Bearer Token kiritilmagan! Iltimos, Bearer Tokenni kiriting."
        }), 400
    
    if not start_date or not end_date:
        return jsonify({
            'success': False,
            'error': "Boshlanish va tugash sanasini tanlang!"
        }), 400

    clean_token = token.strip()
    if clean_token.lower().startswith('bearer '):
        clean_token = clean_token[7:].strip()

    clean_csrf = csrf_token.strip() if csrf_token else "86de5ba7-734b-4b43-93a9-6ebcd5bfaa40"

    if not cookie_str:
        cookie_str = f"account-metadata-Token=eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJleHRyYSI6eyJhdXRoTWV0aG9kIjoicGFzc3dvcmQifSwiYWNjb3VudCI6IjhmMzllYTA4LTA1ZDAtNDdhMC04MTc0LWQ3MWRjYzJmZGE3MSJ9.qbudkhZarEQDmBmWkEJFmsSFr4GEtGlaEZZeGHWWCao; account-metadata-Token.sig=Y7tJBOmOnuE-uA8MAH42cMgNM74; XSRF-TOKEN={clean_csrf}"

    api_url = "https://railway-admin.axonlogic.uz/api/v4/query/admin/orders/download/excel"
    
    headers = {
        'Accept': 'application/json',
        'Accept-Encoding': 'gzip, deflate, br, zstd',
        'Accept-Language': 'ru',
        'Authorization': f"Bearer {clean_token}",
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        'Cookie': cookie_str,
        'Device-Type': 'BROWSER',
        'Host': 'railway-admin.axonlogic.uz',
        'Origin': 'https://railway-admin.axonlogic.uz',
        'Pragma': 'no-cache',
        'Referer': 'https://railway-admin.axonlogic.uz/cabinet/orders',
        'Sec-Ch-Ua': '"Not;A=Brand";v="8", "Chromium";v="150", "Google Chrome";v="150"',
        'Sec-Ch-Ua-Mobile': '?0',
        'Sec-Ch-Ua-Platform': '"macOS"',
        'Sec-Fetch-Dest': 'empty',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Site': 'same-origin',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/150.0.0.0 Safari/150.0.0.0',
        'x-xsrf-token': clean_csrf
    }

    date_from_str = f"{start_date} 00:00:00 +00:00" if len(start_date) == 10 else start_date
    date_to_str = f"{end_date} 23:59:59 +00:00" if len(end_date) == 10 else end_date

    payload = {
        "filterData": {
            "statuses": [
                "ORDER_FINISHED_WITH_EXPRESS_E_TICKET_REGISTRATION_SUCCEEDED"
            ],
            "dateFrom": date_from_str,
            "dateTo": date_to_str
        }
    }

    try:
        resp = requests.post(api_url, json=payload, headers=headers, timeout=120)
        
        if resp.status_code == 200:
            if len(resp.content) < 100 and b"error" in resp.content.lower():
                return jsonify({
                    'success': False,
                    'error': f"API dan kutilmagan javob keldi: {resp.text}",
                    'request_payload': payload
                }), 400

            excel_bytes = None
            try:
                resp_json = resp.json()
                if isinstance(resp_json, dict) and 'data' in resp_json:
                    b64_str = resp_json['data']
                    excel_bytes = base64.b64decode(b64_str)
            except Exception:
                pass

            if not excel_bytes:
                excel_bytes = resp.content

            data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
            report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')

            with open(data_path, 'wb') as f:
                f.write(excel_bytes)

            stats = process_excel(data_path, report_path, uploaded_path=save_path)
            add_upload_log(f"API Sync ({start_date} - {end_date})", len(excel_bytes), "Muvaffaqiyatli")

            return jsonify({
                'success': True,
                'message': f"API orqali {start_date} - {end_date} davri uchun Excel ma'lumotlari muvaffaqiyatli yuklandi va yangilandi!",
                'stats': stats,
                'request_payload': payload
            })
        elif resp.status_code == 504:
            add_upload_log(f"API Sync ({start_date} - {end_date})", 0, "504 Gateway Time-out")
            return jsonify({
                'success': False,
                'error': "504 Gateway Time-out: Railway Admin serveri Excel tayyorlashda taym-aut berdi. Qaytadan urinib ko'ring yoki davrni qisqartiring.",
                'request_payload': payload
            }), 504
        elif resp.status_code == 401:
            add_upload_log(f"API Sync ({start_date} - {end_date})", 0, "401 Unauthorized")
            return jsonify({
                'success': False,
                'error': "401 Unauthorized: Bearer Token yoki Cookie muddati tugagan. Iltimos, yangi tokenlarni kiriting.",
                'request_payload': payload
            }), 401
        elif resp.status_code == 403:
            add_upload_log(f"API Sync ({start_date} - {end_date})", 0, "403 Forbidden")
            return jsonify({
                'success': False,
                'error': f"403 Forbidden: {resp.text[:300]}",
                'request_payload': payload
            }), 403
        else:
            add_upload_log(f"API Sync ({start_date} - {end_date})", 0, f"Xatolik: {resp.status_code}")
            return jsonify({
                'success': False,
                'error': f"API serverida xatolik ({resp.status_code}): {resp.text[:300]}",
                'request_payload': payload
            }), resp.status_code

    except requests.exceptions.Timeout:
        add_upload_log(f"API Sync ({start_date} - {end_date})", 0, "Timeout xatoligi")
        return jsonify({
            'success': False,
            'error': "API so'rovi vaqti tugadi (Timeout 120s).",
            'request_payload': payload
        }), 504
    except Exception as ex:
        add_upload_log(f"API Sync ({start_date} - {end_date})", 0, f"Xatolik: {str(ex)}")
        return jsonify({
            'success': False,
            'error': f"API ga ulanishda xatolik yuz berdi: {str(ex)}"
        }), 500

@app.route('/api/director-summary', methods=['GET'])
def get_director_summary():
    report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
    data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
    try:
        stats = process_excel(data_path, report_path, uploaded_path=save_path)
        return jsonify({
            'success': True,
            'director_summary': stats.get('director_summary', {}),
            'top_stations': stats.get('stations', [])[:5]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/ping', methods=['GET', 'HEAD'])
@app.route('/healthz', methods=['GET', 'HEAD'])
@app.route('/api/ping', methods=['GET', 'HEAD'])
def ping_healthcheck():
    return jsonify({
        'status': 'ok',
        'message': 'pong',
        'server_time': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }), 200

def start_self_ping():
    render_url = os.environ.get('RENDER_EXTERNAL_URL')
    if not render_url:
        return
    
    ping_target = f"{render_url.rstrip('/')}/ping"
    print(f"[Keep-Alive] Self-ping active. Target: {ping_target}")

    def ping_worker():
        import urllib.request
        time.sleep(30)
        while True:
            try:
                req = urllib.request.Request(ping_target, headers={'User-Agent': 'Render-Self-Ping/1.0'})
                with urllib.request.urlopen(req, timeout=10) as resp:
                    if resp.status == 200:
                        print(f"[Keep-Alive] Self-ping successful: {datetime.now().strftime('%H:%M:%S')}")
            except Exception as e:
                print(f"[Keep-Alive] Self-ping warning: {e}")
            time.sleep(600)

    thread = threading.Thread(target=ping_worker, daemon=True)
    thread.start()

def warmup_stats_cache():
    try:
        report_path = os.path.join(app.config['UPLOAD_FOLDER'], 'Август кисока.xlsx')
        data_path = os.path.join(app.config['UPLOAD_FOLDER'], 'data.xlsx')
        global STATS_CACHE
        STATS_CACHE = process_excel(data_path, report_path)
        print("[Cache] Stats cache pre-warmed successfully!")
    except Exception as e:
        print("[Cache] Pre-warmup warning:", e)

threading.Thread(target=warmup_stats_cache, daemon=True).start()
start_self_ping()

def open_browser():
    time.sleep(0.5)
    url = 'http://127.0.0.1:5050'
    try:
        if sys.platform == 'darwin':
            os.system(f'open "{url}"')
        elif sys.platform == 'win32':
            os.system(f'start "" "{url}"')
        else:
            webbrowser.open(url)
    except Exception:
        webbrowser.open(url)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    print("==========================================================")
    print(" Kiosk Hisobot Adminka Senior Web Dasturi ishga tushdi!")
    print(f" Manzil: http://127.0.0.1:{port}")
    print("==========================================================")
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(host='0.0.0.0', port=port, debug=False)

