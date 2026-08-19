import os
import openpyxl
import pandas as pd
from datetime import datetime, date

# 1. Email to station column mapping in 'Худудлар' sheet
# In 'Худудлар' row 2:
# Col D(4), E(5): Андижон
# Col F(6), G(7): Бухоро
# Col H(8), I(9): Хива
# Col J(10), K(11): Тошкент Жанубий
# Col L(12), M(13): Марғилон
# Col N(14), O(15): Навои
# Col P(16), Q(17): Наманган
# Col R(18), S(19): Нукус
# Col T(20), U(21): Қарши
# Col V(22), W(23): Қўнғирод
# Col X(24), Y(25): Қўқон
# Col Z(26), AA(27): Самарқанд
# Col AB(28), AC(29): Термиз
# Col AD(30), AE(31): Тошкент Марказий
# Col AF(32), AG(33): Урганч

EMAIL_MAP = {
    'andijonkiosk@railway.uz': ('Андижон', 4, 5),
    'buxorokiosk@railway.uz': ('Бухоро', 6, 7),
    'khivakiosk@railway.uz': ('Хива', 8, 9),
    'kiosk@axonlogic.uz': ('Тошкент Жанубий', 10, 11),
    'margilonkiosk@railway.uz': ('Марғилон', 12, 13),
    'navoiykiosk@railway.uz': ('Навои', 14, 15),
    'namangankiosk@railway.uz': ('Наманган', 16, 17),
    'nukuskiosk@railway.uz': ('Нукус', 18, 19),
    'qarshikiosk@railway.uz': ('Қарши', 20, 21),
    'qongirotkiosk@railway.uz': ('Қўнғирод', 22, 23),
    'qoqonkiosk@railway.uz': ('Қўқон', 24, 25),
    'samarqandkiosk@railway.uz': ('Самарқанд', 26, 27),
    'termizkiosk@railway.uz': ('Термиз', 28, 29),
    'toshkent.shimoliykiosk@railway.uz': ('Тошкент Марказий', 30, 31),
    'urganchkiosk@railway.uz': ('Урганч', 32, 33)
}

def update_kiosk_report(data_excel_path='data.xlsx', report_excel_path='Август кисока.xlsx'):
    print(f"Reading raw transactions from {data_excel_path}...")
    df = pd.read_excel(data_excel_path)
    
    df['Дата создания_dt'] = pd.to_datetime(df['Дата создания'], errors='coerce')
    df['Date'] = df['Дата создания_dt'].dt.date
    
    kiosk_df = df[df['Пользователь'].isin(EMAIL_MAP.keys())].copy()
    
    ONLINE_PAYMENTS = ['HamkorbankHold', 'HamkorbankWebView', 'Payme', 'StripeIntegration', 'OctoBankFC']
    kiosk_df['PaymentType'] = kiosk_df['Способ оплаты'].apply(
        lambda x: 'Online' if x in ONLINE_PAYMENTS else 'Terminal'
    )

    # Group by Date and Email to get total tickets & total amount
    grouped = kiosk_df.groupby(['Date', 'Пользователь']).agg(
        tickets=('Количество билетов', 'sum'),
        summa=('Общая стоимость', 'sum')
    ).reset_index()

    # Group by Date and PaymentType
    grouped_pay = kiosk_df.groupby(['Date', 'PaymentType']).agg(
        tickets=('Количество билетов', 'sum'),
        summa=('Общая стоимость', 'sum')
    ).reset_index()
    
    # Ensure report file is a valid zip/xlsx file, restore from excellar backup if corrupt
    import zipfile
    if not os.path.exists(report_excel_path) or not zipfile.is_zipfile(report_excel_path):
        ex_backup = os.path.join(os.path.dirname(os.path.abspath(report_excel_path)), 'excellar', os.path.basename(report_excel_path))
        if os.path.exists(ex_backup) and zipfile.is_zipfile(ex_backup):
            import shutil
            shutil.copy(ex_backup, report_excel_path)
            print(f"Restored corrupted {report_excel_path} from excellar backup.")

    print("Loading workbook...")
    wb = openpyxl.load_workbook(report_excel_path)
    
    ws_hudud = wb['Худудлар']
    ws_jami = wb['Жами']
    
    # Build date to row index mapping for 'Худудлар' (Rows 4..34 for Aug 1..31)
    date_to_row = {}
    for r in range(4, 35):
        val = ws_hudud.cell(r, 1).value
        if isinstance(val, (datetime, date)):
            d_obj = val.date() if isinstance(val, datetime) else val
            date_to_row[d_obj] = r

    print("Populating daily station numbers in 'Худудлар' and 'Жами'...")
    # For each date present in grouped data
    for d in sorted(kiosk_df['Date'].dropna().unique()):
        if d in date_to_row:
            r = date_to_row[d]
            day_df = grouped[grouped['Date'] == d]
            
            # First clear row station cells (cols 4 to 33) for this date to avoid leftover stale values
            for c in range(4, 34):
                ws_hudud.cell(r, c).value = None
            
            for email, (st_name, col_soni, col_summa) in EMAIL_MAP.items():
                match = day_df[day_df['Пользователь'] == email]
                if not match.empty:
                    t_val = int(match['tickets'].values[0])
                    s_val = int(match['summa'].values[0])
                    ws_hudud.cell(r, col_soni).value = t_val
                    ws_hudud.cell(r, col_summa).value = s_val
                else:
                    ws_hudud.cell(r, col_soni).value = 0
                    ws_hudud.cell(r, col_summa).value = 0

            # Update 'Жами' sheet payment split
            on_match = grouped_pay[(grouped_pay['Date'] == d) & (grouped_pay['PaymentType'] == 'Online')]
            term_match = grouped_pay[(grouped_pay['Date'] == d) & (grouped_pay['PaymentType'] == 'Terminal')]

            on_t = int(on_match['tickets'].values[0]) if not on_match.empty else 0
            on_s = int(on_match['summa'].values[0]) if not on_match.empty else 0
            term_t = int(term_match['tickets'].values[0]) if not term_match.empty else 0
            term_s = int(term_match['summa'].values[0]) if not term_match.empty else 0

            ws_jami.cell(r, 2).value = f"=Худудлар!B{r}"
            ws_jami.cell(r, 3).value = f"=Худудлар!C{r}"
            ws_jami.cell(r, 4).value = f"=B{r}-F{r}"
            ws_jami.cell(r, 5).value = f"=C{r}-G{r}"
            ws_jami.cell(r, 6).value = term_t
            ws_jami.cell(r, 7).value = term_s

    # Ensure formulas in 'Худудлар' row 35 sum up correctly B4:B34, C4:C34, D4:D34 ... AG4:AG34
    ws_hudud.cell(35, 2).value = "=SUM(B4:B34)"
    ws_hudud.cell(35, 3).value = "=SUM(C4:C34)"
    for c in range(4, 34):
        col_letter = openpyxl.utils.get_column_letter(c)
        ws_hudud.cell(35, c).value = f"=SUM({col_letter}4:{col_letter}34)"

    # Ensure 'Жами' row 35 sums B4:B34, C4:C34
    ws_jami.cell(35, 2).value = "=SUM(B4:B34)"
    ws_jami.cell(35, 3).value = "=SUM(C4:C34)"
    ws_jami.cell(35, 4).value = "=SUM(D4:D34)"
    ws_jami.cell(35, 5).value = "=SUM(E4:E34)"
    ws_jami.cell(35, 6).value = "=SUM(F4:F34)"
    ws_jami.cell(35, 7).value = "=SUM(G4:G34)"

    # Update 'Лист1' if present
    if 'Лист1' in wb.sheetnames:
        ws_list1 = wb['Лист1']
        station_sums = []
        for email, (st_name, col_soni, col_summa) in EMAIL_MAP.items():
            col_soni_letter = openpyxl.utils.get_column_letter(col_soni)
            col_summa_letter = openpyxl.utils.get_column_letter(col_summa)
            
            soni_val = 0
            summa_val = 0
            for r in range(4, 35):
                v_soni = ws_hudud.cell(r, col_soni).value
                v_summa = ws_hudud.cell(r, col_summa).value
                soni_val += int(v_soni) if isinstance(v_soni, (int, float)) else 0
                summa_val += int(v_summa) if isinstance(v_summa, (int, float)) else 0
                
            station_sums.append({
                'stansiya': st_name,
                'soni_val': soni_val,
                'summa_val': summa_val,
                'soni_formula': f"=Худудлар!{col_soni_letter}35",
                'summa_formula': f"=Худудлар!{col_summa_letter}35"
            })

        # Table 1: Sorted by Summa descending
        by_summa = sorted(station_sums, key=lambda x: x['summa_val'], reverse=True)
        for i, item in enumerate(by_summa, start=4):
            ws_list1.cell(i, 1).value = i - 3
            ws_list1.cell(i, 2).value = item['stansiya']
            ws_list1.cell(i, 3).value = item['soni_formula']
            ws_list1.cell(i, 4).value = item['summa_formula']
        ws_list1.cell(20, 3).value = "=SUM(C4:C18)"
        ws_list1.cell(20, 4).value = "=SUM(D4:D18)"

        # Table 2: Sorted by Soni descending
        by_soni = sorted(station_sums, key=lambda x: x['soni_val'], reverse=True)
        for i, item in enumerate(by_soni, start=27):
            ws_list1.cell(i, 1).value = i - 26
            ws_list1.cell(i, 2).value = item['stansiya']
            ws_list1.cell(i, 3).value = item['soni_formula']
            ws_list1.cell(i, 4).value = item['summa_formula']
        ws_list1.cell(43, 3).value = "=SUM(C27:C41)"
        ws_list1.cell(43, 4).value = "=SUM(D27:D41)"

    # Save updated file to both root and excellar/
    wb.save(report_excel_path)
    ex_file_path = os.path.join(os.path.dirname(os.path.abspath(report_excel_path)), 'excellar', os.path.basename(report_excel_path))
    if os.path.exists(os.path.dirname(ex_file_path)):
        wb.save(ex_file_path)
    print("Saved updated workbook with new data and formulas to both root and excellar.")

if __name__ == "__main__":
    update_kiosk_report()

